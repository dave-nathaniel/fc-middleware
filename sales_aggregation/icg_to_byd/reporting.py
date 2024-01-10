import os, sys
import json
import logging
from dotenv import load_dotenv
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from decimal import Decimal
from datetime import datetime
from django.core.mail import EmailMessage

load_dotenv()

def get_log_file():
	logger = logging.getLogger()
	file_handler = None
	for handler in logger.handlers:
		if isinstance(handler, logging.FileHandler):
			file_handler = handler
			break

	return file_handler.baseFilename if file_handler else False


def generate_excel_report(records, file_path):
	# Get the current timestamp and format the timestamp with underscores
	formatted_timestamp = datetime.now().strftime("%Y_%m_%d_%H%M%S")

	# Define headers for the Excel sheet
	headers = ["Date", "Store", "Gross Total", "Net Sales", "VAT", "Consumption Tax", "Tourism Dev. Levy", "Marketing FundProvision", "Locality Marketing Provision", "Mgmt Fee", "Mgmt Fee Share Service", "Mgmt Fee Development", "Mgmt Fee HR", "Rent", "Posted to BYD"]
	
	data = []

	for sale in records:
		data.append([
			sale.date,
			sale.store.store_name,
			sale.gross_total,
			sale.calculated_sales_data.get("net_sales", 0.00),
			sale.calculated_sales_data.get("vat", 0.00),
			sale.calculated_sales_data.get("consumption_tax", 0.00),
			sale.calculated_sales_data.get("tourism_development_levy", 0.00),
			sale.calculated_sales_data.get("marketing_fund_provision", 0.00),
			sale.calculated_sales_data.get("locality_marketing_provision", 0.00),
			sale.calculated_sales_data.get("mgmt_fee", 0.00),
			sale.calculated_sales_data.get("mgmt_fee_share_service", 0.00),
			sale.calculated_sales_data.get("mgmt_fee_development", 0.00),
			sale.calculated_sales_data.get("mgmt_fee_hr", 0.00),
			sale.calculated_sales_data.get("variable_rent", 0.00),
			sale.posted_to_byd,
		])

	df = pd.DataFrame(data, columns=headers)

	# Create a new Excel workbook and add the DataFrame to a new sheet
	wb = openpyxl.Workbook()
	ws = wb.active

	# Add headers
	for col_num, header in enumerate(headers, 1):
		ws.cell(row=1, column=col_num, value=header)

	# Add data
	for r_idx, row in enumerate(data, 2):
		for col_num, value in enumerate(row, 1):
			ws.cell(row=r_idx, column=col_num, value=value)
	
	generated_report = f"reports/{file_path}_{formatted_timestamp}.xlsx"
	# Save the workbook to the specified file path
	wb.save(generated_report)

	return generated_report


def send_email_report(attachment, **kwargs):

	def change_log_to_txt(file_path):
		log_file_path = file_path.split(".")
		log_file_path[-1] = "txt" if log_file_path[-1] == "log" else log_file_path[-1]
		log_file_path = ".".join(log_file_path)

		return log_file_path

	date = kwargs.get('date') if kwargs.get('date', '') != '' else datetime.now()
	date = date.strftime("%d %b %Y")

	active_stores = len(kwargs.get('active_stores')) if kwargs.get('active_stores') else None
	synced_sales = len(kwargs.get('synced_sales')) if kwargs.get('synced_sales') else None
	missing_sales = kwargs.get('missing_sales') if kwargs.get('missing_sales') else []
	posting_errors = kwargs.get('posting_errors') if kwargs.get('posting_errors') else []

	sender_name = f'{os.getenv("EMAIL_SENDER")} for {date}'
	email_from = os.getenv("EMAIL_USER")
	email_to = json.loads(os.getenv("SALES_AGGREGATION_RESULT_EMAIL"))
	email_subject = f"Sales Aggregation Service"
	email_body = ""

	content_subtype = 'html'

	template_file = os.getenv("SALES_AGGREGATION_EMAIL_TEMPLATE")

	try:
		with open(template_file, 'r', encoding='utf-8') as template:
			content = template.read()
	except FileNotFoundError as e:
		logging.error(f"Template file not found in {template_file}")
		return False
	except Exception as e:
		logging.error(f"Template file exception {e}")
		return False

	#Insert the message into the template
	content = content.replace("{{__SALES_AGGREGATION_DATE__}}", str(date))
	content = content.replace("{{__ACTIVE_STORES__}}", str(active_stores))
	content = content.replace("{{__POSTED_SALES__}}", str(synced_sales))
	content = content.replace("{{__NO_SALES_DATA__}}", str(len(missing_sales)))
	content = content.replace("{{__FAILED_POSTING__}}", str(len(posting_errors)))

	missing_sales_table = ""
	for item in missing_sales:
		missing_sales_table += f'<tr><td>{item.store_name}</td></tr>'
	content = content.replace("{{__NO_SALES_DATA_LINE_ITEMS__}}", missing_sales_table)

	posting_errors_table = ""
	for item in posting_errors:
		posting_errors_table += f'<tr><td>{item.store_name}</td></tr>'
	content = content.replace("{{__FAILED_POSTING_LINE_ITEMS__}}", posting_errors_table)

	email_body = content

	email = EmailMessage(
		subject=email_subject, 
		body=email_body, 
		from_email=f"{sender_name} <{email_from}>", 
		to=email_to
	)

	logging.debug(f"{sender_name} <{email_from}>")

	email.content_subtype = content_subtype

	attachments = [attachment, get_log_file()]

	for attachment in attachments:
		if attachment:
			attachment_name = attachment.split("\\")[-1]
			logging.info(f"Adding {attachment_name} as attachment to EMAIL...")
			try:
				# get the file name from the path
				file_name = change_log_to_txt(os.path.basename(attachment))
				with open(attachment, 'rb') as f:
					# add the file as an attachment to the email
					email.attach(file_name, f.read())

			except Exception as e:
				logging.error(e)
				return False

	email.send()

	return True


if __name__ == '__main__':

	import django
	django.setup()

	from store_services.models import Store, Sales

	generate_excel_report([], 'output')